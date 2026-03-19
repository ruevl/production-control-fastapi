import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from src.celery_app import celery_app
from src.core.config import settings
from src.db.session import sync_engine
from src.models.batch import Batch
from src.models.product import Product
from src.models.work_center import WorkCenter
from src.storage.minio_service import get_minio_service


@celery_app.task(bind=True, max_retries=3)
def generate_batch_report(
        self,
        batch_id: int,
        format: str = "excel",
        user_email: str | None = None
):
    try:
        with Session(sync_engine) as db:
            batch = db.execute(
                select(Batch)
                .join(WorkCenter)
                .where(Batch.id == batch_id)
            ).scalar_one_or_none()

            if not batch:
                raise ValueError(f"Batch {batch_id} not found")

            products = db.execute(
                select(Product).where(Product.batch_id == batch_id)
            ).scalars().all()

        with tempfile.NamedTemporaryFile(suffix=f".{format}", delete=False) as tmp_file:
            tmp_path = tmp_file.name

        try:
            if format == "excel":
                _generate_excel_report(tmp_path, batch, products)
            elif format == "pdf":
                _generate_pdf_report(tmp_path, batch, products)
            else:
                raise ValueError(f"Unsupported format: {format}")

            file_size = Path(tmp_path).stat().st_size

            minio_service = get_minio_service()
            object_name = f"batch_{batch_id}_report.{format}"
            file_url = minio_service.upload_file(
                bucket=settings.minio_bucket_reports,
                file_path=tmp_path,
                object_name=object_name,
            )

            return {
                "success": True,
                "file_url": file_url,
                "file_name": object_name,
                "file_size": file_size,
                "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
            }

        finally:
            Path(tmp_path).unlink(missing_ok=True)

    except Exception as exc:
        self.retry(exc=exc, countdown=60)


def _generate_excel_report(file_path: str, batch, products):
    from openpyxl import Workbook

    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "Информация о партии"
    ws_info["A1"] = "Номер партии"
    ws_info["B1"] = batch.batch_number
    ws_info["A2"] = "Дата партии"
    ws_info["B2"] = str(batch.batch_date)
    ws_info["A3"] = "Статус"
    ws_info["B3"] = "Закрыта" if batch.is_closed else "Открыта"
    ws_info["A4"] = "Рабочий центр"
    ws_info["B4"] = batch.work_center.name if batch.work_center else "N/A"
    ws_info["A5"] = "Смена"
    ws_info["B5"] = batch.shift
    ws_info["A6"] = "Бригада"
    ws_info["B6"] = batch.team
    ws_info["A7"] = "Номенклатура"
    ws_info["B7"] = batch.nomenclature
    ws_info["A8"] = "Начало смены"
    ws_info["B8"] = str(batch.shift_start)
    ws_info["A9"] = "Окончание смены"
    ws_info["B9"] = str(batch.shift_end)

    ws_products = wb.create_sheet("Продукция")
    ws_products["A1"] = "ID"
    ws_products["B1"] = "Уникальный код"
    ws_products["C1"] = "Агрегирована"
    ws_products["D1"] = "Дата агрегации"

    for idx, product in enumerate(products, start=2):
        ws_products[f"A{idx}"] = product.id
        ws_products[f"B{idx}"] = product.unique_code
        ws_products[f"C{idx}"] = "Да" if product.is_aggregated else "Нет"
        ws_products[f"D{idx}"] = str(product.aggregated_at) if product.aggregated_at else "-"

    ws_stats = wb.create_sheet("Статистика")
    total = len(products)
    aggregated = sum(1 for p in products if p.is_aggregated)
    ws_stats["A1"] = "Всего продукции"
    ws_stats["B1"] = total
    ws_stats["A2"] = "Агрегировано"
    ws_stats["B2"] = aggregated
    ws_stats["A3"] = "Осталось"
    ws_stats["B3"] = total - aggregated
    ws_stats["A4"] = "Процент выполнения"
    ws_stats["B4"] = f"{(aggregated / total * 100):.2f}%" if total > 0 else "0%"

    wb.save(file_path)


def _generate_pdf_report(file_path: str, batch, products):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 20)
    c.drawString(2 * cm, height - 2 * cm, f"Отчет по партии {batch.batch_number}")

    c.setFont("Helvetica", 12)
    y_pos = height - 4 * cm
    c.drawString(2 * cm, y_pos, f"Дата: {batch.batch_date}")
    y_pos -= 0.5 * cm
    c.drawString(2 * cm, y_pos, f"Рабочий центр: {batch.work_center.name if batch.work_center else 'N/A'}")
    y_pos -= 0.5 * cm
    c.drawString(2 * cm, y_pos, f"Смена: {batch.shift}")
    y_pos -= 0.5 * cm
    c.drawString(2 * cm, y_pos, f"Бригада: {batch.team}")
    y_pos -= 0.5 * cm
    c.drawString(2 * cm, y_pos, f"Номенклатура: {batch.nomenclature}")

    y_pos -= 1 * cm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, y_pos, "Продукция:")

    c.setFont("Helvetica", 10)
    y_pos -= 0.7 * cm
    c.drawString(2 * cm, y_pos, "Код")
    c.drawString(6 * cm, y_pos, "Статус")
    c.drawString(10 * cm, y_pos, "Дата агрегации")

    y_pos -= 0.5 * cm
    for product in products[:50]:
        if y_pos < 2 * cm:
            c.showPage()
            y_pos = height - 2 * cm
        c.drawString(2 * cm, y_pos, product.unique_code)
        c.drawString(6 * cm, y_pos, "Агрегирована" if product.is_aggregated else "Не агрегирована")
        c.drawString(10 * cm, y_pos, str(product.aggregated_at) if product.aggregated_at else "-")
        y_pos -= 0.5 * cm

    c.save()